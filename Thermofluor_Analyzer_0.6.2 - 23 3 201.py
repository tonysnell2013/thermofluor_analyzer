class Thermofluor_Analyzer:
    # Dependencies: openpyxl, matplotlib, xlrd
    
    def __init__(self, filename = '1.xlsx',
                 sheet_name = 'Multicomponent Data',
                 temperature_range = [10, 95]):
        self.filename = filename
        self.sheet_name = sheet_name
        self.temp_range = temperature_range
        
        self.Tm = {}
        self.data = {}
        self.diff = {}
        self.data_norm = {}
        self.diff_norm = {}


    def FileReader(self, filename, sheet_name = "Multicomponent Data"):
        if (filename[-3:] == "xls"):
            return self.FileReader_xls(filename, sheet_name)
        elif(filename[-4:] == "xlsx"):
            return self.FileReader_xlsx(filename, sheet_name)    
        else:
            print ("File Format not supported!")
            return False


    def FileReader_xls(self, filename, sheet_name):
        try:
            import xlrd
            self.wb = xlrd.open_workbook(filename, 'r', on_demand = True)
            self.ws = self.wb.sheet_by_name(sheet_name)

            if (not self.wb.sheet_loaded(sheet_name)):
                print("Sheet Cannot Open!")
                return False

            for irow in range(self.ws.nrows):
                try:
                    row = []
                    for i in range(self.ws.ncols):
                        row.append(self.ws.cell(irow, i).value)

                    if ((isinstance(row[1], (int))
                         or isinstance(row[2], (int, float)))):
                        if ((irow >= 110) or (row[0] in list(self.data.keys()))):
                            self.data[row[0]].append(row[2])
                        else:
                            self.data[row[0]] = [row[2]]
                except:
                    continue

            self.wb.unload_sheet(sheet_name)
            self.wb.release_resources()
            return True

        except ImportError:
            print("Import Error! xlrd not installed!")
            return False
        except FileNotFoundError:
            print ("Error! File not found! Please check the file name / directory")
            return False
        except:
            print("File Read Error!")
            return False

        
    def FileReader_xlsx(self, filename, sheet_name):
        try:
            import openpyxl as opx

            self.wb = opx.load_workbook(filename, read_only = True)
            self.ws = self.wb[sheet_name]
        
            for irow in self.ws.rows:
                if (isinstance(irow[1].value, (int))):
                    if (isinstance(irow[2].value, (int, float))):
                        d = irow[2].value
                    else:
                        d = 0
    
                    if((irow[0].column > 110)
                       or (irow[0].value in list(self.data.keys()))):
                        self.data[irow[0].value].append(d)
                    else:
                        self.data[irow[0].value] = [d]
                else:
                    continue
    
            self.wb._archive.close()
            return True

        except ImportError:
            print("Import Error! openpyxl not installed!")
            return False
        except FileNotFoundError:
            print ("Error! File not found! Please check the file name / directory")
            return False
        except:
            print("File Read Error!")
            return False

    def Analyst(self):
        #try:
        if(True):
            self.data_s = {}
            data_key_sorted = self.SortKey(list(self.data.keys()))
            for comp in data_key_sorted:
                self.diff[comp] = []
                self.data_s[comp] = self.Ave_Smoothen(self.data[comp], 5)

                if(self.Line(self.data_s[comp])):
                    self.Tm[comp] = 0
                    continue

                total = len(self.data_s[comp])
                for i in range(total - 1):
                    self.diff[comp].append(self.data_s[comp][i + 1]
                                           - self.data_s[comp][i])

                self.Tm[comp] = self.Tm_Finder(self.diff[comp])
            return True
        '''
        except ZeroDivisionError:
            print("Error! Division by zero!")
        except:
            print("Analysis Error!")
        '''


    def Ave_Smoothen(self, data_set, n):
        # Smoothen the curve by taking the average of n points

        if((n < 0) or not isinstance(n, int)):
            return "ERROR"
        smoothened = []
        total = len(data_set) - n
        for i in range(total):
            Sum_of_n = 0
            for j in range(n):
                Sum_of_n += data_set[i + j]
            Ave_of_n = Sum_of_n / n
            smoothened.append(Ave_of_n)
        return smoothened


    def Line(self, data_set):
        from scipy import stats

        x = [i for i in range(len(data_set))]
        [g, i, r, p, stderr] = stats.linregress(x, data_set)

        r = r**2
        if(r >= 0.95) and (g <= 0):
            return True
        else:
            return False
        

    def Tm_Finder(self, diff_set):  # Accuracy: +- 1 degree Celsius
        from numpy import polyfit

        max_len = len(diff_set)
        Tm_read = max(diff_set)
        if (Tm_read <= 0):
            return 0

        Ind = diff_set.index(Tm_read)

        sampling_range = 5
        if (Ind >= 0.75 * max_len): # Case 1 -- Tm >= 79
            test_set = diff_set[Ind - sampling_range : Ind]
            for tester in test_set:
                if (abs(tester - Tm_read) >= 0.5 * Tm_read):
                    # Outlier found
                    diff_set[Ind] = 0
                    return self.Tm_Finder(diff_set)
                else: # Not an outlier -- neglegible!
                    [lower, upper] = self.Peak_Range_Finder(diff_set, Ind)
                    for i in range(lower-1, upper):
                        diff_set[i] = 0
                    return self.Tm_Finder(diff_set)

                        
        if (Ind <= 0.1 * max_len): # Case 2 -- Tm <= 18
            test_set = diff_set[Ind : Ind + sampling_range]
            for tester in test_set:
                if (abs(tester - Tm_read) >= 0.5 * Tm_read):
                    # Outlier found
                    diff_set[Ind] = 0
                    return self.Tm_Finder(diff_set)
                else: # Not an outlier -- neglegible!
                    [lower, upper] = self.Peak_Range_Finder(diff_set, Ind)
                    for i in range(lower-1, upper):
                        diff_set[i] = 0
                    return self.Tm_Finder(diff_set)
                    
        [lower, higher] = self.Peak_80Range_Finder(diff_set, Ind)
        x = [i for i in range (lower - 1, higher)]
        fit_dat_set = diff_set[lower - 1 : higher]
        fit_result = polyfit(x, fit_dat_set, 2)
            
        Ind = -fit_result[1]/ (2*fit_result[0])

        if (Ind >= 0.8 * max_len) or (Ind <= 0.1* max_len):
            return 0
        else:
            return (Ind / max_len
                            * (self.temp_range[1] - self.temp_range[0])
                            + self.temp_range[0])
        
    def Plotter(self, Data_Of_Interest = 'A1'):
        
        self.Temp_data = [x / len(self.data[Data_Of_Interest])
                          * (self.temp_range[1] - self.temp_range[0])
                          + self.temp_range[0]
                          for x in range(len(self.data[Data_Of_Interest]))]

        self.Temp_diff = [x / len(self.data[Data_Of_Interest])
                          * (self.temp_range[1] - self.temp_range[0])
                          + self.temp_range[0]
                          for x in range(len(self.diff[Data_Of_Interest]))]


    def Reporter(self, OutFile = 'Output.xlsx'):
        try:
            import openpyxl as opx
        
            self.owb = opx.Workbook(write_only=True)
            self.ows = self.owb.create_sheet('Melting Temperature')

            tem = ["",1,2,3,4,5,6,7,8,9,10,11,12]  # Draw the first line in the output sheet
            self.ows.append(tem)

            Tm_key = self.SortKey(self.Tm.keys())

            len_Tm_key = len(Tm_key)

            for i in range(8):
                new_row = ["%s" % chr(ord('A')+i)] # Draw the alphabetical row number
                for j in range(12):
                    temp_key = "%s%s" % (chr(ord('A') + i), str(j+1))
                    if (temp_key in Tm_key):
                        cell = opx.writer.write_only.WriteOnlyCell(self.ows,
                                                        value = self.Tm[temp_key])
                        new_row.append(cell)
                    else:
                        cell = opx.writer.write_only.WriteOnlyCell(self.ows,
                                                                   value = "")
                        new_row.append(cell)
                self.ows.append(new_row)
            
            if (OutFile[-4:] == "xlsx"):
                self.owb.save(OutFile)
                print("File: \"%s\" is saved." % OutFile)
            else:
                OutFile = "%s.xlsx" % OutFile
                self.owb.save(OutFile)
                print("File: \"%s\" is saved." % OutFile)
            return True

        except ImportError:
            print("Import Error! openpyxl not installed!")
            return False
        except FileNotFoundError:
            print("Error! File not found! Please check the file name / directory")
            return False
        except PermissionError:
            print("Permission denied: \'%s\'" % OutFile)
            return False
        except:
            print("Write File Error!")
            return False
        
    def Raw_Data(self, OutFile = 'Raw_Data_Report.xlsx'): #Not Finished
        import openpyxl as opx

        self.owb = opx.Workbook(write_only=True)
        
        data_key_sort = self.SortKey(self.data.keys())
        for key in data_key_sort:
            # Create a new sheet for each well
            self.ows = self.owb.create_sheet(key)
            # Calculate the plotting coordinates
            self.Plotter(key)   

            # Write coordinates to the file
            self.ows.append(['Temperature'] + self.Temp_data)
            self.ows.append(['Data'] + self.data[key])
            self.ows.append(['Temperature'] + self.Temp_diff)
            self.ows.append(['Derivative'] + self.diff[key])

            # Plot the data
            chart = opx.chart.LineChart()
            

        self.owb.save(OutFile)

    def Peak_Range_Finder(self, data_set, peak_index):
        i = peak_index
        while((data_set[i] > 0) and (i > 0)):
            i -= 1
        lower_bound = i
        i = peak_index
        while((data_set[i] > 0) and (i < len(data_set) - 1)):
            i += 1
        upper_bound = i
        return [lower_bound, upper_bound]

    def Peak_80Range_Finder(self, data_set, peak_index):
        i = peak_index
        Peak = data_set[peak_index]
        cutoff = 0.8 * Peak
        while((data_set[i] > cutoff) and (i > 0)):
            i -= 1
        lower_bound = i
        i = peak_index
        while((data_set[i] > cutoff) and (i < len(data_set) - 1)):
            i += 1
        upper_bound = i
        return [lower_bound, upper_bound]

    def SortKey(self, key_iter):
        # Return a sorted list of keys of a dictionary using the key iterator
        # of such dictionary
    
        key_dict = {}
        for key in key_iter:
            key_dict[key] = ord(key[0].lower()) * 100 + int(key[1:])
        return sorted(key_dict, key=key_dict.__getitem__)


collection = ["20160711_sp1683buffer.xls",
              "20160712_sp0092buffer.xls",
              "20160712_sp0845buffer.xls",
              "20160801 sp0092 sugar.xls",
              "20160803 sp0845 (cut) maltose trial.xls",
              "20160803 sp0845 (cut) sugar.xls",
              "20160809 sp1683 sp1690 sailic acid.xls"]
collection = ["20160804 sp0845 (cut) sugar.xls"]
for Protein in collection:
    x = Thermofluor_Analyzer(filename = Protein)
    if (x.FileReader(Protein)):
        if (x.Analyst()):
            if(x.Reporter("%sOut" % Protein)):
                print("Analysis Complete!")
                
    else:
        print("Program Terminated With Error.")

